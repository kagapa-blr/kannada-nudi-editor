import csv
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QPushButton, \
    QMessageBox, QAction, QLabel
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from utils.asciitounicode import process_line


class ExcelCsvViewer(QMainWindow):
    def __init__(self):
        super().__init__()

        self.file_path = ""
        self.is_modified = False

        self.setWindowTitle("Excel/CSV Viewer - Untitled")
        self.setGeometry(100, 100, 800, 600)

        # Set window icon
        self.setWindowIcon(QIcon('resources/images/logo.jpg'))  # Change this to the path of your parent window logo

        # Set up the layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Label to display the file path
        self.file_path_label = QLabel("")
        self.layout.addWidget(self.file_path_label)

        # Add table to display file data
        self.table = QTableWidget()
        self.layout.addWidget(self.table)

        # Connect table cell change to a slot
        self.table.itemChanged.connect(self.table_item_changed)

        # Create menu bar
        self.create_menu_bar()

    def create_menu_bar(self):
        # Create menu bar
        menu_bar = self.menuBar()

        # Add file menu
        file_menu = menu_bar.addMenu("File")

        # Add open action
        open_action = QAction("Open", self)
        open_action.triggered.connect(self.open_file_prompt_save)
        file_menu.addAction(open_action)

        # Add save action
        save_action = QAction("Save", self)
        save_action.triggered.connect(self.save_file)
        file_menu.addAction(save_action)

    def choose_encoding(self):
        if not self.prompt_save_if_modified():
            return

        # Show a dialog to choose between ASCII and Unicode
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Choose Encoding")
        msg_box.setText("Choose the encoding for the file:")
        ascii_button = msg_box.addButton("ASCII", QMessageBox.ActionRole)
        unicode_button = msg_box.addButton("Unicode", QMessageBox.ActionRole)
        cancel_button = msg_box.addButton(QMessageBox.Cancel)
        msg_box.exec_()

        # Determine the choice
        if msg_box.clickedButton() == ascii_button:
            self.open_file(use_ascii=True)
        elif msg_box.clickedButton() == unicode_button:
            self.open_file(use_ascii=False)
        elif msg_box.clickedButton() == cancel_button:
            pass

    def open_file_prompt_save(self):
        if not self.prompt_save_if_modified():
            return
        self.choose_encoding()

    def open_file(self, use_ascii):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel/CSV File", "",
                                                   "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
                                                   options=options)
        if file_name:
            self.file_path = file_name
            self.file_path_label.setText(f"Opened File: {file_name}")
            self.setWindowTitle(f"Excel/CSV Viewer - {file_name}")

            try:
                if file_name.endswith('.xlsx'):
                    self.load_excel(file_name, use_ascii)
                elif file_name.endswith('.csv'):
                    self.load_csv(file_name, use_ascii)
                self.is_modified = False
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")

    def load_excel(self, file_name, use_ascii):
        workbook = load_workbook(file_name)
        sheet = workbook.active

        rows = sheet.max_row
        columns = sheet.max_column

        self.table.setRowCount(rows)
        self.table.setColumnCount(columns)

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                if value is not None:
                    unicode_value = process_line(str(value)) if use_ascii else str(value)
                    self.table.setItem(i, j, QTableWidgetItem(unicode_value))

    def load_csv(self, file_name, use_ascii):
        with open(file_name, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            data = list(reader)

        rows = len(data)
        columns = len(data[0]) if rows > 0 else 0

        self.table.setRowCount(rows)
        self.table.setColumnCount(columns)

        for i, row in enumerate(data):
            for j, value in enumerate(row):
                unicode_value = process_line(str(value)) if use_ascii else str(value)
                self.table.setItem(i, j, QTableWidgetItem(unicode_value))

    def save_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save File", "",
                                                   "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
                                                   options=options)
        if file_name:
            try:
                if file_name.endswith('.xlsx'):
                    self.save_excel(file_name)
                elif file_name.endswith('.csv'):
                    self.save_csv(file_name)

                self.file_path = file_name
                self.file_path_label.setText(f"Saved File: {file_name}")
                self.setWindowTitle(f"Excel/CSV Viewer - {file_name}")
                self.is_modified = False
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save file: {str(e)}")

    def save_excel(self, file_name):
        workbook = Workbook()
        sheet = workbook.active

        rows = self.table.rowCount()
        columns = self.table.columnCount()

        for i in range(rows):
            for j in range(columns):
                item = self.table.item(i, j)
                if item is not None:
                    sheet.cell(row=i + 1, column=j + 1, value=item.text())

        workbook.save(file_name)

    def save_csv(self, file_name):
        with open(file_name, mode='w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            rows = self.table.rowCount()
            columns = self.table.columnCount()

            for i in range(rows):
                row_data = []
                for j in range(columns):
                    item = self.table.item(i, j)
                    row_data.append(item.text() if item is not None else '')
                writer.writerow(row_data)

    def table_item_changed(self, item):
        self.is_modified = True
        self.setWindowTitle(
            f"Excel/CSV Viewer - {self.file_path} *" if self.file_path else "Excel/CSV Viewer - Untitled *")

    def closeEvent(self, event):
        if self.is_modified:
            reply = QMessageBox.question(self, 'Unsaved Changes',
                                         "You have unsaved changes. Do you want to save them?",
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
            if reply == QMessageBox.Yes:
                self.save_file()
                event.accept()
            elif reply == QMessageBox.No:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    def prompt_save_if_modified(self):
        if self.is_modified:
            reply = QMessageBox.question(self, 'Unsaved Changes',
                                         "You have unsaved changes. Do you want to save them?",
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
            if reply == QMessageBox.Yes:
                self.save_file()
                return True
            elif reply == QMessageBox.No:
                return True
            else:
                return False
        return True
